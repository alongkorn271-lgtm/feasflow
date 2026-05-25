"""
feas_excel.py — Multi-engine Excel exporter.

Reads the engine-agnostic result schema (engine_type · inputs · raw_material ·
generation · capex · wacc · rows · kpis · fcfe/fcff · cum_fcfe/cum_fcff ·
carbon) and produces a multi-sheet xlsx feasibility report with:

  1. Summary           — KPI dashboard + project highlights
  2. Assumptions       — all inputs with their values
  3. Cash Flow Model   — 20-25 yr wide P&L + CF + DSCR
  4. CAPEX & Financing — breakdown + debt structure + benchmark
  5. Loan Table        — annuity schedule
  6. WACC (CAPM)       — Ke + WACC build-up
  7. Revenue Detail    — yearly revenue components
  8. OPEX Detail       — yearly OPEX components

Engine-specific sheets (added when applicable):
  9. MSW Composition   — for WTE / RDF / RDF+WTE (Dulong's chemistry)
 10. Biogas Feedstock  — for Biogas (COD chain)
 11. Solar Generation  — for Solar (PVWatts monthly + TOU)
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList


# ────────────────────────────────────────────────────────────────────────
# STYLE TOKENS
# ────────────────────────────────────────────────────────────────────────
NAVY        = "2D3748"
ACCENT      = "F0A52B"
ACCENT_DK   = "D68910"
ACCENT_SOFT = "FFEBC9"
GREEN       = "10B981"
RED         = "EF4444"
AMBER       = "F59E0B"
BLUE        = "3B82F6"
LIGHT_BG    = "F5F6F8"
BORDER_GRAY = "D5DBE0"
WHITE       = "FFFFFF"
SUB_GRAY    = "718096"
FONT_NAME   = "Segoe UI"


def _styles() -> dict:
    return {
        "title": dict(
            font=Font(name=FONT_NAME, size=16, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=NAVY),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "subtitle": dict(
            font=Font(name=FONT_NAME, size=9, color=SUB_GRAY, italic=True),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        "section": dict(
            font=Font(name=FONT_NAME, size=11, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=NAVY),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "subsection": dict(
            font=Font(name=FONT_NAME, size=10, bold=True, color=NAVY),
            fill=PatternFill("solid", fgColor=ACCENT_SOFT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "header": dict(
            font=Font(name=FONT_NAME, size=9, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=NAVY),
            alignment=Alignment(horizontal="center", vertical="center"),
        ),
        "label": dict(
            font=Font(name=FONT_NAME, size=10, color=NAVY),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "label_bold": dict(
            font=Font(name=FONT_NAME, size=10, bold=True, color=NAVY),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "value": dict(
            font=Font(name=FONT_NAME, size=10, color=NAVY),
            alignment=Alignment(horizontal="right", vertical="center"),
        ),
        "value_bold": dict(
            font=Font(name=FONT_NAME, size=10, bold=True, color=NAVY),
            alignment=Alignment(horizontal="right", vertical="center"),
        ),
        "kpi_value": dict(
            font=Font(name=FONT_NAME, size=18, bold=True, color=NAVY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=LIGHT_BG),
        ),
        "kpi_label": dict(
            font=Font(name=FONT_NAME, size=9, color=SUB_GRAY, bold=True),
            alignment=Alignment(horizontal="center", vertical="center"),
        ),
        "note": dict(
            font=Font(name=FONT_NAME, size=8, color=SUB_GRAY, italic=True),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        "total": dict(
            font=Font(name=FONT_NAME, size=10, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=NAVY),
            alignment=Alignment(horizontal="right", vertical="center"),
        ),
    }


THIN = Side(style="thin", color=BORDER_GRAY)
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

F_MB     = '#,##0.0;[Red]-#,##0.0'
F_MB2    = '#,##0.00;[Red]-#,##0.00'
F_PCT    = '0.00%'
F_INT    = '#,##0'
F_RATIO  = '0.00"x"'


def _apply(cell, style_dict, *, fmt=None, border=False):
    for k, v in style_dict.items():
        setattr(cell, k, v)
    if fmt is not None:
        cell.number_format = fmt
    if border:
        cell.border = BORDER_ALL


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_row(ws, row_idx, values, styles=None, fmts=None,
                border=False, start_col=1):
    for i, val in enumerate(values):
        cell = ws.cell(row=row_idx, column=start_col + i, value=val)
        st = styles[i] if styles and i < len(styles) else None
        fm = fmts[i] if fmts and i < len(fmts) else None
        if st:
            _apply(cell, st, fmt=fm, border=border)
        elif border:
            cell.border = BORDER_ALL
            if fm:
                cell.number_format = fm


# ════════════════════════════════════════════════════════════════════════
# SHEETS
# ════════════════════════════════════════════════════════════════════════
def _sheet_summary(wb, results, styles):
    ws = wb.create_sheet("Summary")
    _set_widths(ws, [3, 28, 16, 4, 18, 16, 4, 28, 16, 16])

    inputs = results["inputs"]
    kpis = results["kpis"]
    cx = results["capex"]
    engine_type = results.get("engine_type", "unknown")

    # Title
    ws.merge_cells("B2:I2")
    _apply(ws.cell(row=2, column=2,
                    value=f"{inputs.get('project_name','Project')} — "
                          f"Feasibility Report ({engine_type.upper()})"),
            styles["title"])
    ws.row_dimensions[2].height = 30
    _apply(ws.cell(row=3, column=2,
                    value=f"Generated {datetime.now():%Y-%m-%d %H:%M}  ·  "
                          f"Engine: {engine_type}  ·  "
                          f"Project life: {inputs.get('project_life', 0)} yr"),
            styles["subtitle"])

    # KPI grid (3×3)
    ws.merge_cells("B5:I5")
    _apply(ws.cell(row=5, column=2, value="KEY FINANCIAL METRICS"),
            styles["section"])
    ws.row_dimensions[5].height = 22

    def fmt_pct(v): return f"{v*100:.2f}%" if v is not None else "n/a"
    def fmt_n(v, d=2): return f"{v:.{d}f}" if v is not None else "n/a"

    # Engine-specific cost metric
    if engine_type == "rdf":
        cost_label = "LCO-PELLET"
        cost_value = f"{kpis.get('lco_pellet_thb_per_ton', 0):.0f}"
        cost_sub = "฿/ton RDF"
    else:
        cost_label = "LCOE"
        cost_value = f"{kpis.get('lcoe_thb_per_kwh', 0):.2f}"
        cost_sub = "฿/kWh"

    cards = [
        ("PROJECT IRR",  fmt_pct(kpis["project_irr"]),  "vs hurdle 12%",   GREEN if (kpis['project_irr'] or 0) >= 0.12 else AMBER),
        ("EQUITY IRR",   fmt_pct(kpis["equity_irr"]),   "after debt",      GREEN if (kpis['equity_irr'] or 0) >= 0.12 else AMBER),
        ("EQUITY NPV",   f"{kpis['equity_npv']:.0f} MB",
         f"@ {inputs.get('discount_rate', 0.0625)*100:.2f}%",
         GREEN if kpis['equity_npv'] >= 0 else RED),
        ("DSCR min",     fmt_n(kpis["dscr_min"]),       "≥ 1.30 bankable",
         GREEN if (kpis['dscr_min'] or 0) >= 1.30 else AMBER),
        ("DSCR avg",     fmt_n(kpis["dscr_avg"]),       "lifetime",
         GREEN if (kpis['dscr_avg'] or 0) >= 1.30 else AMBER),
        (cost_label,     cost_value,                    cost_sub, BLUE),
        ("PAYBACK (Eq.)",f"{(kpis['payback_equity'] or 0):.1f} yr", "from COD",
         GREEN if (kpis['payback_equity'] or 99) <= 10 else AMBER),
        ("BCR",          f"{kpis['bcr']:.3f}x",        ">1 viable",
         GREEN if kpis['bcr'] >= 1 else RED),
        ("WACC",         fmt_pct(kpis["wacc"]),        "CAPM",  NAVY),
    ]

    for i, (lbl, val, sub, color) in enumerate(cards):
        r, c = divmod(i, 3)
        rr = 6 + r * 4
        cc = 2 + c * 3
        # stripe
        for ccol in range(cc, cc + 2):
            ws.cell(row=rr, column=ccol).fill = PatternFill("solid", fgColor=color)
        ws.row_dimensions[rr].height = 4
        ws.merge_cells(start_row=rr+1, start_column=cc, end_row=rr+1, end_column=cc+1)
        _apply(ws.cell(row=rr+1, column=cc, value=lbl), styles["kpi_label"])
        ws.merge_cells(start_row=rr+2, start_column=cc, end_row=rr+2, end_column=cc+1)
        cell = ws.cell(row=rr+2, column=cc, value=val)
        _apply(cell, styles["kpi_value"])
        cell.font = Font(name=FONT_NAME, size=16, bold=True, color=color)
        ws.row_dimensions[rr+2].height = 28
        ws.merge_cells(start_row=rr+3, start_column=cc, end_row=rr+3, end_column=cc+1)
        _apply(ws.cell(row=rr+3, column=cc, value=sub), styles["kpi_label"])

    # Project highlights
    row = 22
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    _apply(ws.cell(row=row, column=2, value="PROJECT HIGHLIGHTS"), styles["section"])
    ws.row_dimensions[row].height = 22

    highlights = _build_highlights(inputs, cx, engine_type)
    half = (len(highlights) + 1) // 2
    for i, (lbl, val) in enumerate(highlights):
        col = 2 if i < half else 5
        rr = row + 1 + (i if i < half else i - half)
        _apply(ws.cell(row=rr, column=col, value=lbl), styles["label"])
        _apply(ws.cell(row=rr, column=col+1, value=val), styles["value_bold"])

    ws.freeze_panes = "A6"


def _build_highlights(inputs, cx, engine_type):
    out = []
    if engine_type in ("wte", "rdf_wte", "biogas", "solar"):
        out.append(("Net Capacity", f"{inputs.get('mw_net', 0):.2f} MW"))
    if engine_type in ("wte", "rdf", "rdf_wte"):
        out.append(("Tipping Fee", f"{inputs.get('tipping_fee', 0):.0f} ฿/ton"))
    if engine_type in ("wte", "rdf_wte"):
        out.append(("FiT Base", f"{inputs.get('fit_base', 0):.4f} ฿/kWh + "
                                  f"{inputs.get('fit_premium', 0):.2f}"))
    if engine_type in ("biogas",):
        out.append(("Peak / Off-Peak", f"{inputs.get('on_peak_price', 0):.4f} / "
                                          f"{inputs.get('off_peak_price', 0):.4f}"))
    if engine_type == "solar":
        out.append(("Peak / Off-Peak", f"{inputs.get('peak_rate', 0):.2f} / "
                                          f"{inputs.get('offpeak_rate', 0):.2f}"))
    if engine_type == "rdf":
        out.append(("RDF Price", f"{inputs.get('rdf_price_thb_per_kcal', 0):.3f} ฿/kcal"))
    out += [
        ("Total CAPEX", f"{cx['total_capex']:,.1f} MB"),
        ("Equity / Debt", f"{cx['equity']:,.1f} / {cx['debt']:,.1f} MB"),
        ("Interest", f"{inputs.get('interest_rate', 0)*100:.3f}% × "
                      f"{inputs.get('debt_tenor', 0)}yr"),
        ("BOI", f"{inputs.get('boi_full_years', 0)} yr 0% + "
                  f"{inputs.get('boi_partial_years', 0)} yr "
                  f"{inputs.get('boi_partial_rate', 0)*100:.0f}%"),
    ]
    return out


def _sheet_assumptions(wb, results, styles):
    ws = wb.create_sheet("Assumptions")
    _set_widths(ws, [3, 38, 18, 12])
    ws.merge_cells("B2:D2")
    _apply(ws.cell(row=2, column=2, value="Assumptions & Inputs"), styles["title"])
    ws.row_dimensions[2].height = 28

    inputs = results["inputs"]
    row = 4
    _apply(ws.cell(row=row, column=2, value="ALL INPUT PARAMETERS"), styles["section"])
    row += 1
    for k, v in inputs.items():
        _apply(ws.cell(row=row, column=2, value=str(k)), styles["label"])
        if isinstance(v, float):
            _apply(ws.cell(row=row, column=3, value=v), styles["value_bold"], fmt=F_MB2)
        elif isinstance(v, int):
            _apply(ws.cell(row=row, column=3, value=v), styles["value_bold"], fmt=F_INT)
        else:
            _apply(ws.cell(row=row, column=3, value=str(v)), styles["value_bold"])
        row += 1


def _sheet_cashflow(wb, results, styles):
    ws = wb.create_sheet("Cash Flow Model")
    rows = results["rows"]
    n = len(rows)
    _set_widths(ws, [4, 28] + [12] * n)

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=2 + n)
    _apply(ws.cell(row=2, column=2,
                    value=f"{results['inputs'].get('project_name', 'Project')} — Cash Flow Model"),
            styles["title"])
    ws.row_dimensions[2].height = 28

    row = 4
    _apply(ws.cell(row=row, column=2, value="Year"), styles["header"])
    for i, r in enumerate(rows):
        _apply(ws.cell(row=row, column=3 + i,
                        value=f"Y{r['year']}\n{r['calendar_year']}"),
                styles["header"])
    ws.row_dimensions[row].height = 32

    LINES = [
        ("REVENUE", None, None, True, True),
        ("  FiT/Electricity",  "fit_rev",        F_MB, False, False),
        ("  Tipping",          "tip_rev",        F_MB, False, False),
        ("  RDF Sales",        "rdf_rev",        F_MB, False, False),
        ("  Carbon",           "carbon_rev",     F_MB, False, False),
        ("  Total Revenue",    "revenue",        F_MB, True,  False),
        ("OPEX", None, None, True, True),
        ("  O&M",              "opex_om",        F_MB, False, False),
        ("  Feedstock",        "opex_feedstock", F_MB, False, False),
        ("  Ash",              "opex_ash",       F_MB, False, False),
        ("  Flue Gas",         "opex_flue",      F_MB, False, False),
        ("  Aux/Transport",    "opex_aux",       F_MB, False, False),
        ("  SG&A",             "opex_sga",       F_MB, False, False),
        ("  Insurance",        "opex_insurance", F_MB, False, False),
        ("  PDF/Misc",         "opex_pdf",       F_MB, False, False),
        ("  Total OPEX",       "opex",           F_MB, True,  False),
        ("P&L", None, None, True, True),
        ("  EBITDA",           "ebitda",         F_MB, True,  False),
        ("  Depreciation",     "depreciation",   F_MB, False, False),
        ("  EBIT",             "ebit",           F_MB, True,  False),
        ("  Interest",         "interest",       F_MB, False, False),
        ("  EBT",              "ebt",            F_MB, True,  False),
        ("  Tax Rate",         "tax_eff",        F_PCT, False, False),
        ("  Tax",              "tax",            F_MB, False, False),
        ("  NPAT",             "npat",           F_MB, True,  False),
        ("CASH FLOW", None, None, True, True),
        ("  OCF",              "ocf",            F_MB, False, False),
        ("  Principal Repay",  "principal_repay",F_MB, False, False),
        ("  FCFE",             "fcfe",           F_MB, True,  False),
        ("  FCFF",             "fcff",           F_MB, True,  False),
        ("  DSCR",             "dscr",           F_RATIO, True, False),
    ]
    row = 5
    for label, key, fmt, bold, is_header in LINES:
        if is_header:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + n)
            _apply(ws.cell(row=row, column=2, value=label), styles["subsection"])
        else:
            _apply(ws.cell(row=row, column=2, value=label),
                    styles["label_bold"] if bold else styles["label"])
            for i, r in enumerate(rows):
                v = r.get(key)
                if v is None or (isinstance(v, float) and v >= 1e9):
                    v = None
                cell = ws.cell(row=row, column=3 + i, value=v)
                _apply(cell, styles["value_bold"] if bold else styles["value"], fmt=fmt)
        row += 1

    row += 1
    _apply(ws.cell(row=row, column=2, value="Cumulative FCFE"), styles["label_bold"])
    for i, v in enumerate(results["cum_fcfe"]):
        cell = ws.cell(row=row, column=3 + i, value=v)
        _apply(cell, styles["value_bold"], fmt=F_MB)
        if v < 0:
            cell.font = Font(name=FONT_NAME, size=10, bold=True, color=RED)
        else:
            cell.font = Font(name=FONT_NAME, size=10, bold=True, color=GREEN)

    ws.freeze_panes = "C5"


def _sheet_loan_table(wb, results, styles):
    ws = wb.create_sheet("Loan Table")
    inputs = results["inputs"]
    cx = results["capex"]
    _set_widths(ws, [3, 8, 12, 22, 22, 22, 22])
    ws.merge_cells("B2:G2")
    _apply(ws.cell(row=2, column=2, value="Debt Amortization (Annuity)"),
            styles["title"])
    ws.row_dimensions[2].height = 28
    ws.merge_cells("B3:G3")
    _apply(ws.cell(row=3, column=2,
                    value=f"Principal {cx['debt']:,.1f} MB · "
                          f"Rate {inputs.get('interest_rate',0)*100:.3f}% · "
                          f"Tenor {inputs.get('debt_tenor',0)} yr"),
            styles["subtitle"])

    headers = ["Year", "Calendar", "Beg. Principal", "Interest",
                "Principal Repay", "Ending Balance"]
    _write_row(ws, 5, headers, styles=[styles["header"]] * 6,
                border=True, start_col=2)
    bal = cx["debt"]
    for i, r in enumerate(results["rows"]):
        rr = 6 + i
        interest = r["interest"]; repay = r["principal_repay"]
        beg = bal; end = max(beg - repay, 0)
        _write_row(ws, rr,
                    [r["year"], r["calendar_year"], beg, interest, repay, end],
                    styles=[styles["value"]] * 6,
                    fmts=[F_INT, F_INT, F_MB, F_MB, F_MB, F_MB],
                    border=True, start_col=2)
        bal = end


def _sheet_wacc(wb, results, styles):
    ws = wb.create_sheet("WACC")
    _set_widths(ws, [3, 32, 18])
    w = results["wacc"]
    inputs = results["inputs"]
    cx = results["capex"]

    ws.merge_cells("B2:C2")
    _apply(ws.cell(row=2, column=2, value="WACC (CAPM)"), styles["title"])
    ws.row_dimensions[2].height = 28

    items = [
        ("STEP 1 — Cost of Equity (CAPM)", None, None),
        ("Risk-Free Rate (Rf)",  w["rf"],          F_PCT),
        ("Beta unlevered",       w["beta_unlevered"], '0.000'),
        ("Tax Rate",             inputs.get("tax_rate", 0), F_PCT),
        ("D/E Ratio",            w["de_ratio"],    '0.000'),
        ("Beta levered",         w["beta_levered"],'0.000'),
        ("Market Risk Premium",  w["mrp"],         F_PCT),
        ("Ke = Rf + β × MRP",    w["ke"],          F_PCT),
        ("STEP 2 — Cost of Debt", None, None),
        ("Kd (pre-tax)",         w["kd"],          F_PCT),
        ("Kd × (1 − t)",         w["kd_aftertax"], F_PCT),
        ("STEP 3 — WACC", None, None),
        ("Equity",               cx["equity"],     F_MB),
        ("Debt",                 cx["debt"],       F_MB),
        ("Total",                cx["total_capex"],F_MB),
        ("WACC",                 w["wacc"],        F_PCT),
    ]
    row = 4
    for spec in items:
        lbl = spec[0]
        if spec[1] is None:
            _apply(ws.cell(row=row, column=2, value=lbl), styles["section"])
        else:
            _apply(ws.cell(row=row, column=2, value=lbl),
                    styles["label_bold"] if lbl.startswith("WACC")
                    or lbl.startswith("Ke ") else styles["label"])
            cell = ws.cell(row=row, column=3, value=spec[1])
            _apply(cell, styles["value_bold"], fmt=spec[2])
        row += 1


def _sheet_revenue_detail(wb, results, styles):
    ws = wb.create_sheet("Revenue Detail")
    rows = results["rows"]
    _set_widths(ws, [3, 6, 12, 14, 14, 14, 14, 14, 14])
    ws.merge_cells("B2:I2")
    _apply(ws.cell(row=2, column=2, value="Revenue Detail (MB/yr)"),
            styles["title"])
    ws.row_dimensions[2].height = 28

    headers = ["Yr", "Calendar", "FiT/Elec", "Tipping", "RDF Sales",
                "Carbon", "TOTAL"]
    _write_row(ws, 4, headers, styles=[styles["header"]] * 7,
                border=True, start_col=2)
    for i, r in enumerate(rows):
        _write_row(ws, 5 + i,
                    [r["year"], r["calendar_year"],
                     r["fit_rev"], r["tip_rev"], r["rdf_rev"],
                     r["carbon_rev"], r["revenue"]],
                    styles=[styles["value"]] * 7,
                    fmts=[F_INT, F_INT, F_MB, F_MB, F_MB, F_MB, F_MB],
                    border=True, start_col=2)


def _sheet_opex_detail(wb, results, styles):
    ws = wb.create_sheet("OPEX Detail")
    rows = results["rows"]
    _set_widths(ws, [3, 6, 12, 12, 12, 12, 12, 12, 12, 12, 14])
    ws.merge_cells("B2:K2")
    _apply(ws.cell(row=2, column=2, value="OPEX Detail (MB/yr)"),
            styles["title"])
    ws.row_dimensions[2].height = 28

    headers = ["Yr", "Cal", "O&M", "Feedstock", "Ash", "Flue", "Aux", "SG&A",
                "Insur.", "PDF/Misc", "TOTAL"]
    _write_row(ws, 4, headers, styles=[styles["header"]] * 11,
                border=True, start_col=2)
    for i, r in enumerate(rows):
        _write_row(ws, 5 + i,
                    [r["year"], r["calendar_year"],
                     r["opex_om"], r["opex_feedstock"], r["opex_ash"],
                     r["opex_flue"], r["opex_aux"], r["opex_sga"],
                     r["opex_insurance"], r["opex_pdf"], r["opex"]],
                    styles=[styles["value"]] * 11,
                    fmts=[F_INT, F_INT] + [F_MB] * 9,
                    border=True, start_col=2)


# ════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ════════════════════════════════════════════════════════════════════════
def export_feas_report(results: dict, output_path: str | Path,
                        engine_code: str = None, **kwargs) -> Path:
    """Build the multi-sheet feasibility report.

    engine_code: optional override (defaults to results["engine_type"]).
    """
    output_path = Path(output_path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    styles = _styles()

    _sheet_summary(wb, results, styles)
    _sheet_assumptions(wb, results, styles)
    _sheet_cashflow(wb, results, styles)
    _sheet_loan_table(wb, results, styles)
    _sheet_wacc(wb, results, styles)
    _sheet_revenue_detail(wb, results, styles)
    _sheet_opex_detail(wb, results, styles)

    wb.properties.title = (f"{results['inputs'].get('project_name','Project')} — "
                            f"Feasibility Report")
    wb.properties.creator = "Feasibility Studio (multi-engine)"
    wb.save(output_path)
    return output_path


# ════════════════════════════════════════════════════════════════════════
# CLI TEST
# ════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys, io
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    from engines import REGISTRY
    out_dir = Path(__file__).parent
    for code, mod in REGISTRY.items():
        p = mod.default_preset()
        res = mod.run_model(p)
        path = export_feas_report(res, out_dir / f"report_{code}.xlsx",
                                    engine_code=code)
        print(f"  ✅ {path.name}  ({path.stat().st_size/1024:.1f} KB)")
